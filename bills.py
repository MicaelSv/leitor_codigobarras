from imbox import Imbox 
from datetime import datetime, timedelta
import pandas as pd
import uuid  # gerador de nomes únicos
from leitor import *
from openpyxl import Workbook
import os


username = 'micael.msilva012@gmail.com'
password = open('passwords/pass', 'r').read() 
host = "imap.gmail.com"
pasta_download = "boletos"

mail = Imbox(host, username = username, password = password, ssl = True)

mensagens = mail.messages(date__gt=datetime.today() - timedelta(days=20), raw='has:attachment') 

wb = Workbook()
ws = wb.active()
r = 1
ws.cell(row=1, column=1).value = "Assunto"
ws.cell(row=1, column=2).value = "Código de barras"
ws.cell(row=1, column=3).value = "Linha digitável"
ws.cell(row=1, column=4).value = "Filename"





for (uid, message) in mensagens:   
    if len(message.attachments) > 0:
        for attach in message.attachments:
            att_file = attach.get('filename')
        
            if '.pdf' in att_file:
                download_path = f"{pasta_download}/{att_file}"

                with open(download_path, 'wb') as fp:
                    fp.write(attach['content'].read())

                try:
                    barcode = leitor_cod(download_path)
                    linha_dig = linha_digitavel(barcode)
                except Exception as e:
                    barcode = False
            
                if not barcode:
                    os.remove(download_path)
                
                else: 
                    print(message.subject, '-', barcode)
                    r += 1
                    #gravando no excel
                    ws.cell(row=r, column=1).value = message.subject
                    ws.cell(row=r, column=2).value = barcode
                    ws.cell(row=r, column=3).value = linha_dig
                    ws.cell(row=r, column=4).value = att_file

wb.save("boletos.xlsx")
mail.logout()

